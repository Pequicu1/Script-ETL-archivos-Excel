"""
Author: Iván López Buira
Date: 21/06/2023
Version: 1.0v
"""

import openpyxl
import os
import subprocess
import pyautogui
import time

INFO = [['B142', 'J142'], ['B127', 'J127'], ['B126', 'J126'], ['B143', 'J143'], [
    'B148', 'J148'], ['B162', 'J162'], ['C35', 'C44'], ['C38', 'C47']]


# Funció per Obrir i guardar els arxius d'excel per calcular les fórmules
# Només cal executarla si no s'han obert previament cada arxius.
def ObrirGuardar(file, path):

    # Obrir arxius excel un per un
    print(file)
    file_path = os.path.join(path, file)

    print(file_path)
    subprocess.Popen([file_path], shell=True)
    time.sleep(1.5)
    pyautogui.hotkey("ctrl", "g")
    time.sleep(1)
    pyautogui.hotkey("alt", "f4")
    time.sleep(1)

# Fució per obtenir la fila o columna de l'arxiu desitjat.


def GetInfo(ws, offset_fila: int, info, file_name, col_ini):

    # Get nom
    ws['B' + str(4 + offset_fila)] = file_name

    offset_columna = 0
    for item in info:
        for celda in item:

            cell = chr(ord(col_ini) - offset_columna) + str(4 + offset_fila)

            print(celda, cell)

            ws[str(cell)] = celda.value

            offset_columna += 1

# Funció que itera sobre les fulles del mateix arxiu d'excel i va omplint les files corresponent a aquella empresa


def GetDataFile(wb, wb2, offset_fila: int, file: str):

    file_name = os.path.splitext(file)[0]

    for hoja in range(0, len(wb.sheetnames)):
        # Aqui escribimos
        print(wb.worksheets[hoja].title)
        ws = wb.worksheets[hoja]
        ws2 = None

        if hoja in range(0, 6):
            ws2 = wb2.worksheets[1]
            col_ini = 'K'
        else:
            ws2 = wb2.worksheets[4]
            col_ini = 'L'

        print(INFO[hoja][0], INFO[hoja][1])
        info_needed = ws2[INFO[hoja][0]:INFO[hoja][1]]

        GetInfo(ws, offset_fila, info_needed, file_name, col_ini)


if __name__ == "__main__":

    dir_dades = "Mila"
    dir_result = "Dades Països"

    for files_res in os.listdir(dir_result):

        pais = files_res.split('.')[0]

        wb = openpyxl.load_workbook(os.path.join(dir_result, files_res))

        dir_dades_pais = os.path.join(dir_dades, pais)

        offset_fila = 0

        for file in os.listdir(dir_dades_pais):

            # (DESCOMENTAR) EXECUTAR AQUESTA LINIA DE CODI NOMÉS SI NO S'HAN OBERT ELS FITXERS EXCEL MAI (PRIMERA EXECUCIÓ).
            #ObrirGuardar(file, dir_dades_pais)

            # itera por cada fichero de la carpeta del Pais.
            print(offset_fila)
            if (file[0] != '~'):
                print(file)

                wb2 = openpyxl.load_workbook(
                    filename=os.path.join(dir_dades_pais, file), data_only=True,  keep_vba=True)

                GetDataFile(wb, wb2, offset_fila, file)
                offset_fila += 1

        wb.save(os.path.join(dir_result, files_res))
